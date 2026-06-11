"""GSTR-3B → Excel consolidation writer."""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ── Colour palette (matches sample) ──────────────────────────────────────────
HDR_BLUE   = 'FF1F3864'   # dark navy  – section header bg
HDR_LIGHT  = 'FFD6E4F0'   # light blue – sub-header bg
HDR_MID    = 'FFB8CCE4'   # mid blue   – column header bg
HDR_MONTH  = 'FFDAE3F3'   # very light – month rows bg
TOTAL_BG   = 'FFFFF2CC'   # yellow     – totals row
WHITE      = 'FFFFFFFF'

NUM_FMT = '#,##0.00'

def _fill(hex_col):
    return PatternFill('solid', fgColor=hex_col)

def _font(bold=False, color='FF000000', size=9, name='Arial'):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

MONTHS = ['April', 'May', 'June', 'July', 'August', 'September',
          'October', 'November', 'December', 'January', 'February', 'March']

BORDER = _border()


def _set(ws, row, col, value, bold=False, fill_hex=None, align_h='center',
         align_v='center', wrap=True, num_fmt=None, font_color='FF000000',
         font_size=9):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=font_color, size=font_size)
    cell.alignment = _align(align_h, align_v, wrap)
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if num_fmt:
        cell.number_format = num_fmt
    cell.border = BORDER
    return cell


def _merge(ws, r1, c1, r2, c2, value, bold=False, fill_hex=None,
           align_h='center', wrap=True, font_color='FFFFFFFF', font_size=9):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = _font(bold=bold, color=font_color, size=font_size)
    cell.alignment = _align(align_h, 'center', wrap)
    if fill_hex:
        cell.fill = _fill(fill_hex)
    cell.border = BORDER
    return cell


def _num(ws, row, col, formula_or_val, fill_hex=None):
    """Write a numeric cell (formula or value)."""
    cell = ws.cell(row=row, column=col, value=formula_or_val)
    cell.font = _font(size=9)
    cell.alignment = _align('right', 'center', False)
    cell.number_format = NUM_FMT
    if fill_hex:
        cell.fill = _fill(fill_hex)
    cell.border = BORDER
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 – output  (Section 3.1 / 3.1.1 / 3.2)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_output_sheet(ws, all_data):
    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    for i in range(3, 45):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # Row heights
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 60
    ws.row_dimensions[4].height = 30

    # ── Row 1: Title ──────────────────────────────────────────────────────
    _merge(ws, 1, 2, 1, 43, 'Section A – Outward / Inward Supplies Summary (GSTR-3B)',
           bold=True, fill_hex=HDR_BLUE, font_size=11, font_color='FFFFFFFF')

    # ── Row 2: Section headers ────────────────────────────────────────────
    sections = [
        (3, 7,  '3.1 Details of Outward supplies and inward supplies liable to reverse charge'),
        (8, 12, '(b) Zero-Rated Supplies'),
        (13, 17,'(c) Nil / Exempt Supplies'),
        (18, 22,'(d) RCM Inward'),
        (23, 27,'(e) Non-GST Outward'),
        (28, 32,'3.1.1 (i) ECO Sec 9(5) – ECO pays'),
        (33, 37,'3.1.1 (ii) ECO Sec 9(5) – Registered Person'),
        (38, 39,'3.2 – Unregistered'),
        (40, 41,'3.2 – Composition'),
        (42, 43,'3.2 – UIN Holders'),
    ]
    for c1, c2, label in sections:
        color = HDR_MID if (c1 - 3) % 10 < 5 else HDR_LIGHT
        _merge(ws, 2, c1, 2, c2, label, bold=True, fill_hex=color,
               font_color='FF1F3864', font_size=8)

    # ── Row 3: Sub-headers ────────────────────────────────────────────────
    _set(ws, 3, 2, 'Month', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    sub_cols = ['Taxable Value', 'IGST', 'CGST', 'SGST', 'Cess']
    # sections (a)-(e) and 3.1.1(i)&(ii)
    for grp in range(7):
        for j, sub in enumerate(sub_cols):
            col = 3 + grp * 5 + j
            _set(ws, 3, col, sub, bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=8)

    # 3.2 columns (only Taxable Value + IGST)
    for grp, label in enumerate(['Unreg', 'Comp', 'UIN']):
        _set(ws, 3, 38 + grp*2, 'Taxable Value', bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=8)
        _set(ws, 3, 39 + grp*2, 'IGST', bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=8)

    # ── Rows 4–15: Month data ─────────────────────────────────────────────
    month_row_map = {}
    for i, month in enumerate(MONTHS):
        row = 4 + i
        month_row_map[month.lower()] = row
        bg = HDR_MONTH if i % 2 == 0 else WHITE
        _set(ws, row, 2, month, bold=False, fill_hex=bg, align_h='left')

        # Write zeros initially; filled after
        for col in range(3, 44):
            _num(ws, row, col, 0, fill_hex=bg)

    # Fill actual data
    for d in all_data:
        month = d['period'].lower()
        row = month_row_map.get(month)
        if not row:
            continue
        bg = HDR_MONTH if MONTHS.index(d['period'].capitalize()) % 2 == 0 else WHITE

        def wr(col_start, vals):
            for j, v in enumerate(vals[:5]):
                _num(ws, row, col_start + j, v, fill_hex=bg)

        wr(3,  d['3_1a'])
        wr(8,  d['3_1b'])
        wr(13, d['3_1c'])
        wr(18, d['3_1d'])
        wr(23, d['3_1e'])
        wr(28, d['3_1_1i'])
        wr(33, d['3_1_1ii'])

        # 3.2
        _num(ws, row, 38, d['3_2_unreg'][0], fill_hex=bg)
        _num(ws, row, 39, d['3_2_unreg'][1], fill_hex=bg)
        _num(ws, row, 40, d['3_2_comp'][0], fill_hex=bg)
        _num(ws, row, 41, d['3_2_comp'][1], fill_hex=bg)
        _num(ws, row, 42, d['3_2_uin'][0], fill_hex=bg)
        _num(ws, row, 43, d['3_2_uin'][1], fill_hex=bg)

    # ── Row 16: Total ─────────────────────────────────────────────────────
    row = 16
    _set(ws, row, 2, 'TOTAL', bold=True, fill_hex=TOTAL_BG, align_h='center', font_color='FF1F3864')
    for col in range(3, 44):
        c_letter = get_column_letter(col)
        _num(ws, row, col, f'=SUM({c_letter}4:{c_letter}15)', fill_hex=TOTAL_BG)

    ws.freeze_panes = 'C4'


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 – ITC (Section 4)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_itc_sheet(ws, all_data):
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    for i in range(3, 45):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 35
    ws.row_dimensions[3].height = 50
    ws.row_dimensions[4].height = 25

    # Row 1 title
    _merge(ws, 1, 2, 1, 43, 'Section B – Eligible ITC (Section 4, GSTR-3B)',
           bold=True, fill_hex=HDR_BLUE, font_size=11, font_color='FFFFFFFF')

    # Row 2 group headers
    itc_groups = [
        (3, 6,   'A(1) Import of Goods'),
        (7, 10,  'A(2) Import of Services'),
        (11, 14, 'A(3) RCM (other than 1&2)'),
        (15, 18, 'A(4) ISD'),
        (19, 22, 'A(5) All Other ITC'),
        (23, 26, 'B(1) Rules 38,42,43 & Sec 17(5)'),
        (27, 30, 'B(2) Others'),
        (31, 34, 'C. Net ITC (A-B)'),
        (35, 38, 'D(1) ITC Reclaimed'),
        (39, 42, 'D(2) Ineligible ITC (16(4)/PoS)'),
    ]
    for c1, c2, label in itc_groups:
        color = HDR_MID if itc_groups.index((c1, c2, label)) % 2 == 0 else HDR_LIGHT
        _merge(ws, 2, c1, 2, c2, label, bold=True, fill_hex=color,
               font_color='FF1F3864', font_size=8)

    # Row 3 sub-headers
    _set(ws, 3, 2, 'Month', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    for grp in range(10):
        for j, sub in enumerate(['IGST', 'CGST', 'SGST', 'Cess']):
            _set(ws, 3, 3 + grp*4 + j, sub, bold=True, fill_hex=HDR_MID,
                 font_color='FF1F3864', font_size=8)

    # Month rows
    month_row_map = {}
    for i, month in enumerate(MONTHS):
        row = 4 + i
        month_row_map[month.lower()] = row
        bg = HDR_MONTH if i % 2 == 0 else WHITE
        _set(ws, row, 2, month, bold=False, fill_hex=bg, align_h='left')
        for col in range(3, 43):
            _num(ws, row, col, 0, fill_hex=bg)

    for d in all_data:
        month = d['period'].lower()
        row = month_row_map.get(month)
        if not row:
            continue
        bg = HDR_MONTH if MONTHS.index(d['period'].capitalize()) % 2 == 0 else WHITE

        def wr(col_start, vals):
            for j, v in enumerate(vals[:4]):
                _num(ws, row, col_start + j, v, fill_hex=bg)

        wr(3,  d['itc_a1'])
        wr(7,  d['itc_a2'])
        wr(11, d['itc_a3'])
        wr(15, d['itc_a4'])
        wr(19, d['itc_a5'])
        wr(23, d['itc_b1'])
        wr(27, d['itc_b2'])
        wr(31, d['itc_c'])
        wr(35, d['itc_d1'])
        wr(39, d['itc_d2'])

    # Total row
    row = 16
    _set(ws, row, 2, 'TOTAL', bold=True, fill_hex=TOTAL_BG, align_h='center', font_color='FF1F3864')
    for col in range(3, 43):
        c_letter = get_column_letter(col)
        _num(ws, row, col, f'=SUM({c_letter}4:{c_letter}15)', fill_hex=TOTAL_BG)

    ws.freeze_panes = 'C4'


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 – Section 5 + 5.1
# ═══════════════════════════════════════════════════════════════════════════════

def _build_sec5_sheet(ws, all_data):
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    for i in range(3, 22):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 35
    ws.row_dimensions[3].height = 30

    _merge(ws, 1, 2, 1, 21, 'Section C – Exempt Inward Supplies & Interest/Late Fee (Sec 5 & 5.1)',
           bold=True, fill_hex=HDR_BLUE, font_size=11, font_color='FFFFFFFF')

    groups = [
        (3, 4,   '5. Composition/Exempt/Nil (Inter/Intra)'),
        (5, 6,   '5. Non-GST Supply (Inter/Intra)'),
        (7, 10,  '5.1 System Computed Interest'),
        (11, 14, '5.1 Interest Paid'),
        (15, 18, '5.1 Late Fee'),
    ]
    for c1, c2, label in groups:
        color = HDR_MID if groups.index((c1, c2, label)) % 2 == 0 else HDR_LIGHT
        _merge(ws, 2, c1, 2, c2, label, bold=True, fill_hex=color,
               font_color='FF1F3864', font_size=8)

    _set(ws, 3, 2, 'Month', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    for col, sub in [(3, 'Inter-State'), (4, 'Intra-State'), (5, 'Inter-State'), (6, 'Intra-State')]:
        _set(ws, 3, col, sub, bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=8)
    for grp_start in [7, 11, 15]:
        for j, sub in enumerate(['IGST', 'CGST', 'SGST', 'Cess']):
            _set(ws, 3, grp_start + j, sub, bold=True, fill_hex=HDR_MID,
                 font_color='FF1F3864', font_size=8)

    month_row_map = {}
    for i, month in enumerate(MONTHS):
        row = 4 + i
        month_row_map[month.lower()] = row
        bg = HDR_MONTH if i % 2 == 0 else WHITE
        _set(ws, row, 2, month, bold=False, fill_hex=bg, align_h='left')
        for col in range(3, 19):
            _num(ws, row, col, 0, fill_hex=bg)

    for d in all_data:
        month = d['period'].lower()
        row = month_row_map.get(month)
        if not row:
            continue
        bg = HDR_MONTH if MONTHS.index(d['period'].capitalize()) % 2 == 0 else WHITE
        _num(ws, row, 3,  d['sec5_comp'][0], fill_hex=bg)
        _num(ws, row, 4,  d['sec5_comp'][1], fill_hex=bg)
        _num(ws, row, 5,  d['sec5_nongst'][0], fill_hex=bg)
        _num(ws, row, 6,  d['sec5_nongst'][1], fill_hex=bg)
        for j, v in enumerate(d['interest_paid']):
            _num(ws, row, 11 + j, v, fill_hex=bg)
        for j, v in enumerate(d['late_fee']):
            _num(ws, row, 15 + j, v, fill_hex=bg)

    row = 16
    _set(ws, row, 2, 'TOTAL', bold=True, fill_hex=TOTAL_BG, align_h='center', font_color='FF1F3864')
    for col in range(3, 19):
        c_letter = get_column_letter(col)
        _num(ws, row, col, f'=SUM({c_letter}4:{c_letter}15)', fill_hex=TOTAL_BG)

    ws.freeze_panes = 'C4'


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 4 – Payment of Tax (6.1)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_payment_sheet(ws, all_data):
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16  # Tax type
    for i in range(4, 35):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 55
    ws.row_dimensions[4].height = 25

    _merge(ws, 1, 2, 1, 33, 'Section D – Payment of Tax (Section 6.1, GSTR-3B)',
           bold=True, fill_hex=HDR_BLUE, font_size=11, font_color='FFFFFFFF')

    # Group headers (row 2)
    # Cols: B=Month, C=Tax Type, then 10 data cols per row × 8 tax types
    # Layout: Col D–M = (A) IGST, N–W = (A) CGST, X–AG = (A) SGST, AH–AQ = (A) Cess
    #         AR–BA = (B) IGST ... etc.
    # Simplified: show each month as a block with 8 sub-rows (tax types)

    # Better: Transpose – rows=months, columns = tax_type × metric
    # 8 tax types × 10 columns = 80 columns – too wide
    # Use rows: Month | Tax Category | Tax Type | 10 metrics
    # Row structure: for each month, 8 sub-rows (A:IGST,CGST,SGST,Cess; B:same)

    # Header structure: row 2 = main metric groups, row 3 = sub-columns, row 4 = data cols
    metric_groups = [
        (4, 4,   'Tax Payable'),
        (5, 5,   'Adj Neg Liability'),
        (6, 6,   'Net Tax Payable'),
        (7, 7,   'ITC – IGST'),
        (8, 8,   'ITC – CGST'),
        (9, 9,   'ITC – SGST'),
        (10, 10, 'ITC – Cess'),
        (11, 11, 'Cash Paid'),
        (12, 12, 'Interest (Cash)'),
        (13, 13, 'Late Fee (Cash)'),
    ]
    _merge(ws, 2, 2, 3, 2, 'Month', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    _merge(ws, 2, 3, 3, 3, 'Tax Category / Type', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')

    for c, (c1, c2, label) in enumerate(metric_groups):
        color = HDR_MID if c % 2 == 0 else HDR_LIGHT
        _set(ws, 2, c1, label, bold=True, fill_hex=color, font_color='FF1F3864', font_size=8)
        ws.row_dimensions[2].height = 50

    # Data rows: for each month, 8 tax-type sub-rows
    # Sections: (A) IGST, (A) CGST, (A) SGST, (A) Cess, (B) IGST, (B) CGST, (B) SGST, (B) Cess
    tax_type_labels = [
        '(A) IGST', '(A) CGST', '(A) SGST', '(A) Cess',
        '(B) IGST', '(B) CGST', '(B) SGST', '(B) Cess',
    ]
    data_keys = ['pay_a_igst', 'pay_a_cgst', 'pay_a_sgst', 'pay_a_cess',
                 'pay_b_igst', 'pay_b_cgst', 'pay_b_sgst', 'pay_b_cess']

    # Build lookup
    month_data = {}
    for d in all_data:
        month_data[d['period'].lower()] = d

    current_row = 4
    for i, month in enumerate(MONTHS):
        start_row = current_row
        bg = HDR_MONTH if i % 2 == 0 else WHITE

        for j, (tax_label, key) in enumerate(zip(tax_type_labels, data_keys)):
            row = current_row + j
            ws.row_dimensions[row].height = 16

            if j == 0:
                # Merge month cell across 8 rows
                ws.merge_cells(start_row=start_row, start_column=2,
                               end_row=start_row + 7, end_column=2)
                cell = ws.cell(row=start_row, column=2, value=month)
                cell.font = _font(bold=True, size=9, color='FF1F3864')
                cell.alignment = _align('center', 'center', False)
                cell.fill = _fill(bg)
                cell.border = BORDER

            _set(ws, row, 3, tax_label, bold=False, fill_hex=bg,
                 align_h='left', font_size=8)

            d = month_data.get(month.lower())
            vals = d[key] if d and key in d else [0]*10
            for k, v in enumerate(vals[:10]):
                _num(ws, row, 4 + k, v, fill_hex=bg)

        current_row += 8

    # Total rows
    _merge(ws, current_row, 2, current_row + 7, 2, 'TOTAL',
           bold=True, fill_hex=TOTAL_BG, font_color='FF1F3864')
    for j, tax_label in enumerate(tax_type_labels):
        row = current_row + j
        _set(ws, row, 3, tax_label, bold=True, fill_hex=TOTAL_BG,
             align_h='left', font_color='FF1F3864', font_size=8)
        # Sum every 8th row for this tax type
        for col in range(4, 14):
            c_letter = get_column_letter(col)
            # Rows for this tax type: 4+j, 12+j, 20+j, ... (every 8)
            refs = ','.join(
                f'{c_letter}{4 + j + m*8}'
                for m in range(12)
                if (4 + j + m*8) < current_row
            )
            _num(ws, row, col, f'=SUM({refs})' if refs else 0, fill_hex=TOTAL_BG)

    ws.freeze_panes = 'D4'


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN: build_excel
# ═══════════════════════════════════════════════════════════════════════════════

def build_excel(all_data, output_path):
    """Build consolidated GSTR-3B Excel workbook."""
    wb = openpyxl.Workbook()

    ws_sum = wb.active
    ws_sum.title = 'SUMMARY'
    ws_out = wb.create_sheet('Outward Supplies (3.1)')
    ws_itc = wb.create_sheet('Eligible ITC (4)')
    ws_s5  = wb.create_sheet('Exempt & Interest (5)')
    ws_pay = wb.create_sheet('Payment of Tax (6.1)')

    _build_summary_sheet(ws_sum, all_data)
    _build_output_sheet(ws_out, all_data)
    _build_itc_sheet(ws_itc, all_data)
    _build_sec5_sheet(ws_s5, all_data)
    _build_payment_sheet(ws_pay, all_data)

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 0 – SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def _build_summary_sheet(ws, all_data):
    """Build SUMMARY sheet with cross-sheet formulas matching sample format."""

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 3
    for col in ['C','D','E','F','G','H','I','J','K','L','M','N','O','P',
                'Q','R','S','T','U','V','W','X','Y']:
        ws.column_dimensions[col].width = 14

    # ── PERIOD table (rows 1-14) ──────────────────────────────────────────────
    _merge(ws, 1, 3, 1, 8, 'PERIOD', bold=True, fill_hex=HDR_BLUE,
           font_size=11, font_color='FFFFFFFF')

    _set(ws, 2, 3, 'year',          bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    _set(ws, 2, 4, 'period(month)', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    _set(ws, 2, 5, 'DATE of ARN',   bold=True, fill_hex=HDR_MID, font_color='FF1F3864')

    # Build lookup: month -> metadata
    month_meta = {}
    for d in all_data:
        month_meta[d['period'].capitalize()] = {
            'year':     d.get('year', '2025-26'),
            'arn':      d.get('arn', ''),
            'arn_date': d.get('arn_date', ''),
            'gstin':    d.get('gstin', ''),
            'state':    d.get('state', ''),
        }

    # Column headers: year | period | DATE of ARN | ARN | GSTIN | STATE
    _set(ws, 2, 6, 'ARN',   bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    _set(ws, 2, 7, 'GSTIN', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    _set(ws, 2, 8, 'STATE', bold=True, fill_hex=HDR_MID, font_color='FF1F3864')
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 18

    for i, month in enumerate(MONTHS):
        row = 3 + i
        bg = HDR_MONTH if i % 2 == 0 else WHITE
        meta = month_meta.get(month, {})
        _set(ws, row, 3, meta.get('year', '2025-26'), fill_hex=bg, align_h='center')
        _set(ws, row, 4, month,                       fill_hex=bg, align_h='center')
        # DATE of ARN
        arn_date_str = meta.get('arn_date', '')
        if arn_date_str:
            from datetime import datetime
            try:
                arn_date_val = datetime.strptime(arn_date_str, '%d/%m/%Y')
            except Exception:
                arn_date_val = arn_date_str
        else:
            arn_date_val = ''
        cell = ws.cell(row=row, column=5, value=arn_date_val)
        cell.font = _font(size=9)
        cell.alignment = _align('center', 'center', False)
        cell.fill = _fill(bg)
        cell.border = BORDER
        if arn_date_val and hasattr(arn_date_val, 'year'):
            cell.number_format = 'DD/MM/YYYY'
        # ARN
        _set(ws, row, 6, meta.get('arn',   ''), fill_hex=bg, align_h='center', font_size=8)
        # GSTIN
        _set(ws, row, 7, meta.get('gstin', ''), fill_hex=bg, align_h='center', font_size=9)
        # STATE
        _set(ws, row, 8, meta.get('state', ''), fill_hex=bg, align_h='left',   font_size=9)

    # ── SECTION-A header (row 15-16) ─────────────────────────────────────────
    _merge(ws, 15, 3, 15, 15, 'SECTION -A', bold=True, fill_hex=HDR_BLUE,
           font_size=10, font_color='FFFFFFFF')

    _merge(ws, 16, 3,  16, 8,  'GSTR-3B :OUTWARD(EXCLUDING RCM)',
           bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=9)
    _merge(ws, 16, 9,  16, 9,  '', fill_hex=WHITE, font_color='FF000000')
    _merge(ws, 16, 10, 16, 15, 'INWARD SUPPLIES LIABLE TO RCM',
           bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=9)

    # ── SECTION-A column headers (row 17) ─────────────────────────────────────
    for col, label in [(3,'Month'),(4,'Taxable Value'),(5,'IGST'),(6,'CGST'),
                       (7,'SGST'),(8,'Cess'),(10,'Month'),(11,'Taxable Value'),
                       (12,'IGST'),(13,'CGST'),(14,'SGST'),(15,'Cess')]:
        _set(ws, 17, col, label, bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=8)
    _set(ws, 17, 9, '', fill_hex=WHITE)

    # ── SECTION-A data rows (rows 18-29) ──────────────────────────────────────
    # Outward Supplies sheet columns (data rows 4-15):
    #   C=3.1a TV, D=3.1a IGST, E=3.1a CGST, F=3.1a SGST, G=3.1a Cess
    #   H=3.1b TV, I=3.1b IGST, J=3.1b CGST, K=3.1b SGST, L=3.1b Cess
    #   M=3.1c TV, N=3.1c IGST, O=3.1c CGST, P=3.1c SGST, Q=3.1c Cess
    #   R=3.1d TV, S=3.1d IGST, T=3.1d CGST, U=3.1d SGST, V=3.1d Cess
    OS = "'Outward Supplies (3.1)'"

    for i, month in enumerate(MONTHS):
        row = 18 + i
        src = 4 + i   # source row in Outward sheet
        bg = HDR_MONTH if i % 2 == 0 else WHITE

        _set(ws, row, 3, month, fill_hex=bg, align_h='left')
        # Outward TV = sum of 3.1a + 3.1b + 3.1c taxable values
        _num(ws, row, 4, f"={OS}!C{src}+{OS}!H{src}+{OS}!M{src}", fill_hex=bg)
        _num(ws, row, 5, f"={OS}!D{src}+{OS}!I{src}+{OS}!N{src}", fill_hex=bg)
        _num(ws, row, 6, f"={OS}!E{src}+{OS}!J{src}+{OS}!O{src}", fill_hex=bg)
        _num(ws, row, 7, f"={OS}!F{src}+{OS}!K{src}+{OS}!P{src}", fill_hex=bg)
        _num(ws, row, 8, f"={OS}!G{src}+{OS}!L{src}+{OS}!Q{src}", fill_hex=bg)

        _set(ws, row, 9, '', fill_hex=WHITE)

        _set(ws, row, 10, month, fill_hex=bg, align_h='left')
        _num(ws, row, 11, f"={OS}!R{src}", fill_hex=bg)
        _num(ws, row, 12, f"={OS}!S{src}", fill_hex=bg)
        _num(ws, row, 13, f"={OS}!T{src}", fill_hex=bg)
        _num(ws, row, 14, f"={OS}!U{src}", fill_hex=bg)
        _num(ws, row, 15, f"={OS}!V{src}", fill_hex=bg)

    # SECTION-A TOTAL row (row 30)
    _set(ws, 30, 3, 'TOTAL', bold=True, fill_hex=TOTAL_BG, align_h='center', font_color='FF1F3864')
    for col in range(4, 9):
        cl = get_column_letter(col)
        _num(ws, 30, col, f"=SUM({cl}18:{cl}29)", fill_hex=TOTAL_BG)
    _set(ws, 30, 9, '', fill_hex=WHITE)
    _set(ws, 30, 10, 'TOTAL', bold=True, fill_hex=TOTAL_BG, align_h='center', font_color='FF1F3864')
    for col in range(11, 16):
        cl = get_column_letter(col)
        _num(ws, 30, col, f"=SUM({cl}18:{cl}29)", fill_hex=TOTAL_BG)

    # ── SECTION-B header (row 31-32) ─────────────────────────────────────────
    _merge(ws, 31, 3, 31, 25, 'SECTION -B', bold=True, fill_hex=HDR_BLUE,
           font_size=10, font_color='FFFFFFFF')
    _merge(ws, 32, 3, 32, 25, 'ELIGIBLE ITC', bold=True, fill_hex=HDR_MID,
           font_color='FF1F3864', font_size=10)

    # ── SECTION-B sub-section headers (row 33) ────────────────────────────────
    _merge(ws, 33, 3,  33, 7,  'ITC :EXCLUDING RCM',
           bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=8)
    _set(ws, 33, 8, '', fill_hex=WHITE)
    _merge(ws, 33, 9,  33, 13, 'ITC: RCM',
           bold=True, fill_hex=HDR_LIGHT, font_color='FF1F3864', font_size=8)
    _set(ws, 33, 14, '', fill_hex=WHITE)
    _merge(ws, 33, 15, 33, 19, 'ITC:ISD',
           bold=True, fill_hex=HDR_MID, font_color='FF1F3864', font_size=8)
    _set(ws, 33, 20, '', fill_hex=WHITE)
    _merge(ws, 33, 21, 33, 25, 'INELIGIBLE ITC',
           bold=True, fill_hex=HDR_LIGHT, font_color='FF1F3864', font_size=8)

    # ── SECTION-B column headers (row 34) ─────────────────────────────────────
    sub = ['Month','IGST','CGST','SGST','Cess']
    for grp_start in [3, 9, 15, 21]:
        for j, lbl in enumerate(sub):
            _set(ws, 34, grp_start + j, lbl, bold=True, fill_hex=HDR_MID,
                 font_color='FF1F3864', font_size=8)
    for blank_col in [8, 14, 20]:
        _set(ws, 34, blank_col, '', fill_hex=WHITE)

    # ── SECTION-B data rows (rows 35-46) ─────────────────────────────────────
    # ITC sheet column mapping (data rows 4-15):
    #   A(1): C=IGST D=CGST E=SGST F=Cess
    #   A(2): G=IGST H=CGST I=SGST J=Cess
    #   A(3)RCM: K=IGST L=CGST M=SGST N=Cess
    #   A(4)ISD: O=IGST P=CGST Q=SGST R=Cess
    #   A(5):    S=IGST T=CGST U=SGST V=Cess
    #   B(1):    W=IGST X=CGST Y=SGST Z=Cess
    #   B(2):    AA=IGST AB=CGST AC=SGST AD=Cess
    #   C Net:   AE=IGST AF=CGST AG=SGST AH=Cess
    #   D(1):    AI=IGST AJ=CGST AK=SGST AL=Cess
    #   D(2):    AM=IGST AN=CGST AO=SGST AP=Cess
    ITC = "'Eligible ITC (4)'"

    for i, month in enumerate(MONTHS):
        row = 35 + i
        src = 4 + i
        bg = HDR_MONTH if i % 2 == 0 else WHITE

        # ITC Excl. RCM = A(1)+A(2)+A(5) - B(1) - B(2)  [per sample formula]
        _set(ws, row, 3, month, fill_hex=bg, align_h='left')
        _num(ws, row, 4, f"={ITC}!C{src}+{ITC}!G{src}+{ITC}!S{src}-{ITC}!W{src}-{ITC}!AA{src}", fill_hex=bg)
        _num(ws, row, 5, f"={ITC}!D{src}+{ITC}!H{src}+{ITC}!T{src}-{ITC}!X{src}-{ITC}!AB{src}", fill_hex=bg)
        _num(ws, row, 6, f"={ITC}!E{src}+{ITC}!I{src}+{ITC}!U{src}-{ITC}!Y{src}-{ITC}!AC{src}", fill_hex=bg)
        _num(ws, row, 7, f"={ITC}!F{src}+{ITC}!J{src}+{ITC}!V{src}-{ITC}!Z{src}-{ITC}!AD{src}", fill_hex=bg)

        _set(ws, row, 8, '', fill_hex=WHITE)

        # ITC RCM = A(3)
        _set(ws, row, 9, month, fill_hex=bg, align_h='left')
        _num(ws, row, 10, f"={ITC}!K{src}", fill_hex=bg)
        _num(ws, row, 11, f"={ITC}!L{src}", fill_hex=bg)
        _num(ws, row, 12, f"={ITC}!M{src}", fill_hex=bg)
        _num(ws, row, 13, f"={ITC}!N{src}", fill_hex=bg)

        _set(ws, row, 14, '', fill_hex=WHITE)

        # ITC ISD = A(4)
        _set(ws, row, 15, month, fill_hex=bg, align_h='left')
        _num(ws, row, 16, f"={ITC}!O{src}", fill_hex=bg)
        _num(ws, row, 17, f"={ITC}!P{src}", fill_hex=bg)
        _num(ws, row, 18, f"={ITC}!Q{src}", fill_hex=bg)
        _num(ws, row, 19, f"={ITC}!R{src}", fill_hex=bg)

        _set(ws, row, 20, '', fill_hex=WHITE)

        # Ineligible ITC = D(2): AM=IGST AN=CGST AO=SGST AP=Cess
        _set(ws, row, 21, month, fill_hex=bg, align_h='left')
        _num(ws, row, 22, f"={ITC}!AM{src}", fill_hex=bg)
        _num(ws, row, 23, f"={ITC}!AN{src}", fill_hex=bg)
        _num(ws, row, 24, f"={ITC}!AO{src}", fill_hex=bg)
        _num(ws, row, 25, f"={ITC}!AP{src}", fill_hex=bg)

    # SECTION-B TOTAL row (row 47)
    for grp_start, grp_end in [(3,7),(9,13),(15,19),(21,25)]:
        _set(ws, 47, grp_start, 'TOTAL', bold=True, fill_hex=TOTAL_BG,
             align_h='center', font_color='FF1F3864')
        for col in range(grp_start+1, grp_end+1):
            cl = get_column_letter(col)
            _num(ws, 47, col, f"=SUM({cl}35:{cl}46)", fill_hex=TOTAL_BG)
    for blank_col in [8, 14, 20]:
        _set(ws, 47, blank_col, '', fill_hex=WHITE)

    ws.freeze_panes = 'D18'
